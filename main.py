from openpyxl import load_workbook
import customtkinter
from tkinter import filedialog
import os


customtkinter.set_appearance_mode("System")
customtkinter.set_default_color_theme("blue")

root = customtkinter.CTk()

# Set the window title font and foreground color
root.title("Beta version v0.2 Leonid")

#root.iconbitmap("C:/Users/Sahar/Desktop/pythonProject4/SEDG.ICO")
root.geometry(f"{1100}x{600}")
root.grid_columnconfigure((0, 1, 2, 3,4,5,6,7), weight=1)
root.grid_rowconfigure((0, 1, 2, 3, 5, 6, 7, 8, 9, 10, 11), weight=1)

# Declare global variables
source_file_paths = [""] * 20
target_file_path = ""  # Initialize target_file_path

def process():
    # Adjust these values based on your requirements
    source_config = {
        'sheet': 'sheet2',
        'start_row': 6,
        'end_row': 47,
        'column_v1': 3,
        'column_v2': 4,
        'step': 2,
    }
    dest_file_path = entry_dest_file_path.get()
    if not dest_file_path:
        customtkinter.messagebox.showwarning("Warning", "Please choose a destination file path.")
        return
    if not os.path.exists(dest_file_path):
        customtkinter.messagebox.showwarning("Warning", f"Destination file path does not exist: {dest_file_path}")
        return

    target_workbook = load_workbook(dest_file_path)  # Initialize target workbook
    source_workbook = None

    for i in range(20):
        entry_path_value = entry_paths[i].get()
        if entry_path_value:
            # Create a new target configuration for each iteration
            target_config = {
                'sheet': 'SOC X%',
                'target_column_v1': 16,
                'target_column_v2': 17,
                'target_row': 27 + i * 26,
            }
            source_workbook = process_file(entry_path_value, source_config, target_workbook, target_config)
            transfer_f6_to_p16(source_workbook, target_workbook, target_config)
    # Close workbooks after the loop
    if source_workbook:
        source_workbook.close()
    if target_workbook:
        updated_target_file_path(target_workbook, dest_file_path)

    root.destroy()

def process_file(source_file, source_config, target_workbook, target_config):
    source_workbook = load_workbook(source_file)

    copy_values_between_files(
        source_file, source_config,
        target_workbook, target_config,
        source_workbook
    )
    return source_workbook

def transfer_f6_to_p16(source_workbook, target_workbook, target_config):
    source_sheet = source_workbook['sheet2']
    target_sheet = target_workbook[target_config['sheet']]

    # Transfer the value from F6 in the source file to P16 in the target file
    source_cell_f6 = source_sheet['F6']
    target_cell_p16 = target_sheet['P16']
    target_cell_p16.value = source_cell_f6.value

def choose_source_file(i):
    global source_file_paths
    file_path = filedialog.askopenfilename(title=f"Select Source File {i + 1}")
    entry_paths[i].delete(0, customtkinter.END)
    entry_paths[i].insert(0, file_path)
    source_file_paths[i] = file_path

# Create entry widgets and buttons for source file paths
entry_paths = []
buttons_choose_source = []

for i in range(20):
    entry_column = 0 if i < 10 else 3
    button_column = 1 if i < 10 else 4
    entry_row = i % 10
    entry_path = customtkinter.CTkEntry(root, placeholder_text=f"Month {i + 1}", width=200, )
    entry_path.grid(row=entry_row, column=entry_column, padx=1, pady=1)
    button_choose_source = customtkinter.CTkButton(root, text=f"Choose Source File {i + 1}",command=lambda i=i: choose_source_file(i))
    button_choose_source.grid(row=entry_row, column=button_column, padx=1, pady=1)
    entry_paths.append(entry_path)
    buttons_choose_source.append(button_choose_source)

button_process = customtkinter.CTkButton(root, text="Process", command=process)
button_process.grid(row=10, column=7, columnspan=2, padx=(20, 20), pady=(20, 20), sticky="nsew")

entry_dest_file_path = customtkinter.CTkEntry(root, width=200)
entry_dest_file_path.grid(row=5, column=7)

def choose_dest_file_path():
    dest_file_path = filedialog.askopenfilename(
        defaultextension=".xlsx",
        filetypes=[("Excel files", "*.xlsx")],
        title="Choose Destination File Path"
    )
    if dest_file_path:
        entry_dest_file_path.delete(0, customtkinter.END)
        entry_dest_file_path.insert(0, dest_file_path)


button_choose_dest_file_path = customtkinter.CTkButton(root, text="Choose Destination File Path", command=choose_dest_file_path)
button_choose_dest_file_path.grid(row=5, column=6,sticky="w")

def copy_values_between_files(source_file, source_config, target_workbook, target_config, source_workbook):
    source_sheet = source_workbook[source_config['sheet']]
    target_sheet = target_workbook[target_config['sheet']]

    for row_index in range(source_config['start_row'], source_config['end_row'] + 1, source_config['step']):
        source_cell_v1 = source_sheet.cell(row=row_index, column=source_config['column_v1'])
        target_cell_v1 = target_sheet.cell(row=target_config['target_row'], column=target_config['target_column_v1'])
        source_cell_v2 = source_sheet.cell(row=row_index, column=source_config['column_v2'])
        target_cell_v2 = target_sheet.cell(row=target_config['target_row'], column=target_config['target_column_v2'])
        target_cell_v1.value = source_cell_v1.value
        target_cell_v2.value = source_cell_v2.value

        target_config['target_row'] += 1

    # Save the target workbook after processing all source files

def generate_updated_filename(original_filepath):
    if not original_filepath:
        return ""  # Return an empty string if the original filepath is empty

    filepath, filename = os.path.split(original_filepath)
    filename, extension = os.path.splitext(filename)
    updated_filename = f"{filename}_Updated{extension}"
    return os.path.join(filepath, updated_filename)


# Function to generate the updated file name
def updated_target_file_path(target_workbook, file_path):
    updated_target_file_path = generate_updated_filename(file_path)
    target_workbook.save(updated_target_file_path)
    target_workbook.close()



# Set target sheet information (SOC X%)
target_sheet = "SOC X%"
target_column_v1 = 16
target_column_v2 = 17
target_row = 27

root.mainloop()
