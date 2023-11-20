from openpyxl import load_workbook
import customtkinter
from tkinter import filedialog
import openpyxl

customtkinter.set_appearance_mode("System")  # Modes: "System" (standard), "Dark", "Light"
customtkinter.set_default_color_theme("blue")  # Themes: "blue" (standard), "green", "dark-blue"

root = customtkinter.CTk()
root.title("SolarEdge")
root.iconbitmap("C:/Users/Sahar/Desktop/leonid_test/SEDG.ICO")
root.geometry(f"{900}x{600}")
root.grid_columnconfigure(( 0, 1, 2, 3,  ), weight=1)
root.grid_rowconfigure((0, 1, 2, 3, 5, 6, 7, 8, 9, 10, 11, ), weight=1)

# Declare global variables
source_file_path = ""
#target_file_path = ""


def process():
    global source_file_path

    # Copy the value between files
    copy_values_v1_between_files(
        source_file_path, source_sheet, start_row, end_row,
        source_column_v1, target_file, target_sheet,
        target_column_v1
    )
    copy_value_between_files(
        source_file_path, source_sheet, target_file, target_sheet, source_cell, target_cell
    )

def choose_source_file():
    global source_file_path
    file_path = filedialog.askopenfilename(title="Select Source File")
    entry_path_1.delete(0, customtkinter.END)
    entry_path_1.insert(0, file_path)
    source_file_path = file_path


def choose_target_file():
    global target_file_path
    file_path = filedialog.askopenfilename(title="Select Target File")
    # Add an entry widget or label to display the selected target file path if needed
    target_entry.delete(0, customtkinter.END)
    target_entry.insert(0, file_path)
    target_file_path = file_path

def copy_value_between_files(source_file_path, source_sheet, target_file, target_sheet,source_cell,   target_cell):
    # Open the source workbook and sheet
    source_workbook = openpyxl.load_workbook(source_file_path)
    source_sheet = source_workbook[source_sheet]

    # Get the value from the source cell
    source_value = source_sheet[source_cell].value

    # Open the target workbook and sheet
    target_workbook = openpyxl.load_workbook(target_file)

    # Check if the target sheet exists, create it if not
    if target_sheet not in target_workbook.sheetnames:
        target_workbook.create_sheet(title=target_sheet)

    target_sheet = target_workbook[target_sheet]

    # Put the value into the target cell
    target_sheet[target_cell].value = source_value

    # Save the changes to the target workbook
    target_workbook.save(target_file)


# Specify the file paths and cell locations

source_sheet_name = "sheet2"  # Adjust the sheet name as needed
source_cell_location = "F6"

target_file_path = "C:/Users/Sahar/Desktop/leonid_test/UpdatedFormatForDCIRRPT.xlsx"
target_sheet_name = "Sheet1"  # Adjust the sheet name as needed
target_cell_location = "P16"

entry_path_1 = customtkinter.CTkEntry(root, width=330)
entry_path_1.grid(row=0, column=0, )

button_choose_source_1 = customtkinter.CTkButton(root, text="Choose Source File", command=choose_source_file)
button_choose_source_1.grid(row=0, column=1, padx=1, pady=1)

entry_path_2 = customtkinter.CTkEntry(root, width=330)
entry_path_2.grid(row=1, column=0, padx=1, pady=1,)

button_choose_source_2 = customtkinter.CTkButton(root, text="Choose Source File", command=choose_source_file)
button_choose_source_2.grid(row=1, column=1, padx=1, pady=1)

entry_path_3 = customtkinter.CTkEntry(root, width=330)
entry_path_3.grid(row=2, column=0, padx=1, pady=1)

button_choose_source_3 = customtkinter.CTkButton(root, text="Choose Source File", command=choose_source_file)
button_choose_source_3.grid(row=2, column=1, padx=1, pady=1)

entry_path_4 = customtkinter.CTkEntry(root, width=330)
entry_path_4.grid(row=3, column=0, padx=1, pady=1, )

button_choose_source_4 = customtkinter.CTkButton(root, text="Choose Source File", command=choose_source_file)
button_choose_source_4.grid(row=3, column=1, padx=1, pady=1)

entry_path_5 = customtkinter.CTkEntry(root, width=330)
entry_path_5.grid(row=4, column=0, padx=1, pady=1, )

button_choose_source_5 = customtkinter.CTkButton(root, text="Choose Source File", command=choose_source_file)
button_choose_source_5.grid(row=4, column=1, padx=1, pady=1)

entry_path_6 = customtkinter.CTkEntry(root, width=330)
entry_path_6.grid(row=5, column=0, padx=1, pady=1, )

button_choose_source_6 = customtkinter.CTkButton(root, text="Choose Source File", command=choose_source_file)
button_choose_source_6.grid(row=5, column=1, padx=1, pady=1)

entry_path_7 = customtkinter.CTkEntry(root, width=330)
entry_path_7.grid(row=6, column=0, padx=1, pady=1,)

button_choose_source_7 = customtkinter.CTkButton(root, text="Choose Source File", command=choose_source_file)
button_choose_source_7.grid(row=6, column=1, padx=1, pady=1)

entry_path_8 = customtkinter.CTkEntry(root, width=330)
entry_path_8.grid(row=7, column=0, padx=1, pady=1,)

button_choose_source_8 = customtkinter.CTkButton(root, text="Choose Source File", command=choose_source_file)
button_choose_source_8.grid(row=7, column=1, padx=1, pady=1)

entry_path_7 = customtkinter.CTkEntry(root, width=330)
entry_path_7.grid(row=8, column=0, padx=1, pady=1, )

button_choose_source_9 = customtkinter.CTkButton(root, text="Choose Source File", command=choose_source_file)
button_choose_source_9.grid(row=8, column=1, padx=1, pady=1)

entry_path_10 = customtkinter.CTkEntry(root, width=330)
entry_path_10.grid(row=9, column=0, padx=1, pady=1)

button_choose_source_10 = customtkinter.CTkButton(root, text="Choose Source File", command=choose_source_file)
button_choose_source_10.grid(row=9, column=1, padx=1, pady=1)


#button_choose_target = customtkinter.CTkButton(root, text="Choose Target File", command=choose_target_file)
#button_choose_target.grid(row=1, column=1, padx=10, pady=10)

# Add an entry widget or label to display the selected target file path if needed
#target_entry = customtkinter.CTkEntry(root, width=350)
#target_entry.grid(row=2, column=0, padx=20, pady=20, columnspan=2)

button_process = customtkinter.CTkButton(root, text="Process File", command=process)
button_process.grid(row=10, column=3, columnspan=2, padx=10, pady=10)


def copy_values_v1_between_files(source_file, source_sheet, start_row, end_row, source_column, target_file,
                                 target_sheet, target_column):
    source_workbook = load_workbook(source_file)
    source_sheet = source_workbook[source_sheet]

    target_workbook = load_workbook(target_file)
    target_sheet = target_workbook[target_sheet]

    target_row = 27  # Start copying to P27

    for row_index in range(start_row, end_row + 1, step):
        source_cell_v1 = source_sheet.cell(row=row_index, column=source_column_v1)
        target_cell_v1 = target_sheet.cell(row=target_row, column=target_column_v1)
        source_cell_v2 = source_sheet.cell(row=row_index, column=source_column_v2)
        target_cell_v2 = target_sheet.cell(row=target_row, column=target_column_v2)
        target_cell_v1.value = source_cell_v1.value
        target_cell_v2.value = source_cell_v2.value

        target_row += 1

    target_workbook.save(target_file)
    source_workbook.close()
    target_workbook.close()



source_sheet = "sheet2"
start_row = 6
end_row = 28
source_column_v1 = 3  # Column C
source_column_v2 = 4
#source_column_I2 = 6
#source_row_I2 = 6
source_cell = "F6"
step = 2
target_file ="C:/Users/Sahar/Desktop/leonid_test/UpdatedFormatForDCIRRPT.xlsx"
target_sheet = "Sheet1"
target_column_v1 = 16  # Column P
target_column_v2 = 17
#target_column_I2 = 16
#target_row_I2 = 16
target_cell="P16"
root.mainloop()